from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
from openpyxl import Workbook


options = webdriver.ChromeOptions()
options.add_argument(("lang=ko_KR"))
options.add_experimental_option('excludeSwitches', ['enable-logging'])

# 브라우저 창 사이즈
options.add_argument('window-size=1280x720')

title_list = []

def crawling():
    try :
        # 드라이버 경로 입력
        s = Service('chromedriver.exe')
        driver = webdriver.Chrome(service=s, options=options)

        driver.get("https://www.filemaru.com/index.php")

        time.sleep(6)

        btn = driver.find_element(By.XPATH, '//*[@id="container_area2"]/div/ul/li[11]/a')
        btn.click()

        time.sleep(2)

        for j in range(2,40047):
            time.sleep(3)
            titles = driver.find_elements(By.CLASS_NAME, "ctit")
            for i in titles:
                text = i.text
                title_list.append(text)
                print(text)
            driver.execute_script(f"pagingOpen('rightListPage','400456','10','5','{j}','');")
        
    finally:
        print(title_list)
        filing()

        

def filing():
    try:
        wb = Workbook()

        ws = wb.active

        ws.title = "새 시트"

        ws['A1'] = "title"
        
        for i, j in enumerate(title_list):
            ws[f'A{i + 2}'] = j

        wb.save('filemaru_protect_list.xlsx')
    finally:
        print("에러")

if __name__ == "__main__":
    crawling()