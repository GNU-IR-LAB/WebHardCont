
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from bs4 import BeautifulSoup as bs
from selenium.webdriver.common.alert import Alert
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
import time
from openpyxl import Workbook



options = webdriver.ChromeOptions()
options.add_argument(("lang=ko_KR"))
options.add_experimental_option('excludeSwitches', ['enable-logging'])
# options.add_extension('bihmplhobchoageeokmgbdihknkjbknd.crx')

# 브라우저 창 사이즈
options.add_argument('window-size=1280x720')

# 드라이버 경로 입력
s = Service('chromedriver.exe')
driver = webdriver.Chrome(service=s, options=options)
wait = WebDriverWait(driver, 10)

driver.get(f'https://www.filekuki.com/')
time.sleep(1)
driver.switch_to.frame('main')
driver.find_element(By.XPATH, '//*[@id="id"]').send_keys('aing3773')
driver.find_element(By.XPATH, '//*[@id="inputPwd"]').send_keys('7uXUrs@NjrmA')
driver.find_element(By.XPATH, '//*[@id="loginSubmit"]').click()
driver.switch_to.default_content()
time.sleep(2)


driver.switch_to.frame('main')
driver.find_element(By.XPATH, '//*[@id="footer"]/div/div[1]/ul/li[7]').click()
driver.find_element(By.XPATH, '//*[@id="rank_movie"]/table/tbody/tr[2]/td/table/tbody/tr/td[6]').click()
driver.switch_to.default_content()

title_list = []

for j in range(11650):
    driver.switch_to.frame('main')
    for i in range(2, 41, 2):
        title = driver.find_element(By.XPATH, f'//*[@id="rank_movie"]/table/tbody/tr[7]/td/table/tbody/tr[{i}]/td[2]').text
        title_list.append(title)
        print(title)
    if j == 1:
        driver.find_element(By.XPATH, '//*[@id="rank_movie"]/table/tbody/tr[8]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[22]').click()
    if j > 20:
        driver.find_element(By.XPATH, '//*[@id="rank_movie"]/table/tbody/tr[8]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[27]').click()
    else:
        driver.find_element(By.XPATH, '//*[@id="rank_movie"]/table/tbody/tr[8]/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[25]').click()

        
    driver.switch_to.default_content()
    time.sleep(1)


wb = Workbook()

ws = wb.active

ws.title = "새 시트"

ws['A1'] = "title"

for i, j in enumerate(title_list):
    ws[f'A{i + 2}'] = j

wb.save('filemaru_protect_list.xlsx')

while True:
    pass